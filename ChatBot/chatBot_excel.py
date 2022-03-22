'''
엑셀로 저장된 질문과 답변을 가져와 txt 파일로 저장한다
-> 이후 말뭉치 모델 학습에 이용
'''
import openpyxl

def make_txt(): # 텍스트 파일 생성 (이름 : data.txt)
    f = open("./data.txt","w")
    f.close()
    return

def save_txt(a): # 텍스트 파일에 작성
    f = open("./data.txt","a")
    f.write(a)
    f.close()
    return

def read_excel(): # 엑셀을 읽어와 텍스트 파일에 저장
    make_txt()
    a =""
    wb = openpyxl.load_workbook('./sample.xlsx')    #예제 엑셀파일을 열어 workbook instance 객체로 저장
    sheet = wb['database'] # workbook 객체에서 worksheet 이름을 index로 사용해 instance 객체를 가지고 온다
    for i in range (1, sheet.max_row+1):
        for j in range (1, sheet.max_column+1):
            a += sheet.cell(row = i, column = j).value + " "
        a+="\n"
        save_txt(a)
        a = ""
    wb.close()
    return

read_excel()
